from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
from datetime import datetime
import os
import subprocess
import tempfile
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from models import db, User, Transaction, Category
from config import Config

app = Flask(__name__)
app.config.from_object(Config)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB
db.init_app(app)
CORS(app)

with app.app_context():
    db.create_all()
    if not User.query.first():
        db.session.add(User(name='Edward', email='edward@example.com'))
        db.session.add(User(name='Bernice', email='bernice@example.com'))
        db.session.commit()
    if not Category.query.first():
        for t, n in [('expense','伙食'),('expense','杂费'),('expense','马票'),('expense','赌博'),('expense','房屋贷款'),('expense','CP 500'),('expense','Side Income'),('income','借贷'),('income','收入'),('income','银行利息/股息'),('income','WL Salary'),('income','HMSB Incentive'),('income','OJ Incentive'),('income','Lepas Incentive'),('income','SleepyFace Studio Account')]:
            if not Category.query.filter_by(type=t, name=n).first():
                db.session.add(Category(type=t, name=n))
        db.session.commit()

@app.route('/api/users', methods=['GET'])
def get_users():
    return jsonify([{'id': u.id, 'name': u.name, 'email': u.email} for u in User.query.all()])

@app.route('/api/transactions', methods=['POST'])
def add_transaction():
    try:
        data = request.get_json()
        t = Transaction(user_id=data.get('user_id'), type=data.get('type'), category=data.get('category'), amount=float(data.get('amount')), date=data.get('date', datetime.now().strftime('%Y-%m-%d')), note=data.get('note', ''))
        db.session.add(t)
        db.session.commit()
        return jsonify({'id': t.id, 'message': '交易记录已保存'}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/transactions', methods=['GET'])
def get_transactions():
    user_id, start_date, end_date = request.args.get('user_id'), request.args.get('start_date'), request.args.get('end_date')
    q = Transaction.query
    if user_id: q = q.filter_by(user_id=user_id)
    if start_date: q = q.filter(Transaction.date >= start_date)
    if end_date: q = q.filter(Transaction.date <= end_date)
    return jsonify([{'id': t.id, 'user_id': t.user_id, 'type': t.type, 'category': t.category, 'amount': t.amount, 'date': t.date, 'note': t.note} for t in q.all()])

@app.route('/api/statistics', methods=['GET'])
def get_statistics():
    user_id = request.args.get('user_id')
    q = Transaction.query
    if user_id: q = q.filter_by(user_id=user_id)
    ts = q.all()
    te = sum(t.amount for t in ts if t.type == 'expense')
    ti = sum(t.amount for t in ts if t.type == 'income')
    return jsonify({'total_expense': te, 'total_income': ti, 'balance': ti - te})

@app.route('/api/categories', methods=['GET'])
def get_categories():
    return jsonify([{'id': c.id, 'type': c.type, 'name': c.name} for c in Category.query.all()])

@app.route('/api/categories', methods=['POST'])
def add_category():
    try:
        data = request.get_json()
        c = Category(type=data.get('type'), name=data.get('name'))
        db.session.add(c)
        db.session.commit()
        return jsonify({'id': c.id, 'message': '分类已保存'}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/import/transactions', methods=['POST'])
def import_transactions():
    """导入Excel文件中的交易数据"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': '没有选择文件'}), 400
        
        file = request.files['file']
        password = request.form.get('password', '')
        user_id = request.form.get('user_id', '')
        
        if not file or file.filename == '':
            return jsonify({'error': '文件为空'}), 400
        
        if not user_id:
            return jsonify({'error': '必须指定用户ID'}), 400
        
        # 验证用户存在
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': '用户不存在'}), 400
        
        # 保存临时文件
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            file.save(tmp.name)
            temp_file = tmp.name
        
        try:
            # 尝试直接打开文件
            wb = load_workbook(temp_file, data_only=True)
        except:
            # 如果失败（可能是加密的），尝试用LibreOffice处理
            if password:
                try:
                    temp_decrypted = temp_file + '.decrypted.xlsx'
                    # LibreOffice命令行处理加密文件
                    subprocess.run([
                        'libreoffice', '--headless', '--convert-to', 'xlsx',
                        '--outdir', os.path.dirname(temp_file),
                        temp_file
                    ], timeout=30, capture_output=True)
                    
                    # 重命名解密后的文件
                    base_name = os.path.basename(temp_file)
                    converted = os.path.join(os.path.dirname(temp_file), base_name.replace('.xlsx', '.xlsx'))
                    if os.path.exists(converted):
                        temp_file = converted
                except:
                    pass
            
            # 再次尝试打开
            try:
                wb = load_workbook(temp_file, data_only=True)
            except Exception as e:
                return jsonify({'error': f'无法打开文件: {str(e)}'}), 400
        
        ws = wb.active
        imported_count = 0
        errors = []
        
        # 获取表头（第一行）
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        # 预期的列映射（根据Excel的列标题）
        column_map = {}
        for i, header in enumerate(headers):
            if header:
                header_lower = str(header).lower().strip()
                if '日期' in header_lower or 'date' in header_lower:
                    column_map['date'] = i
                elif '类型' in header_lower or 'type' in header_lower:
                    column_map['type'] = i
                elif '分类' in header_lower or 'category' in header_lower:
                    column_map['category'] = i
                elif '金额' in header_lower or 'amount' in header_lower:
                    column_map['amount'] = i
                elif '备注' in header_lower or 'note' in header_lower or 'description' in header_lower:
                    column_map['note'] = i
        
        # 如果没有找到表头，使用默认顺序
        if not column_map:
            column_map = {'date': 0, 'type': 1, 'category': 2, 'amount': 3, 'note': 4}
        
        # 从第二行开始读取数据
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row or all(cell is None for cell in row):
                    continue
                
                # 提取字段
                date_val = row[column_map.get('date', 0)] if column_map.get('date') < len(row) else None
                type_val = row[column_map.get('type', 1)] if column_map.get('type') < len(row) else None
                category_val = row[column_map.get('category', 2)] if column_map.get('category') < len(row) else None
                amount_val = row[column_map.get('amount', 3)] if column_map.get('amount') < len(row) else None
                note_val = row[column_map.get('note', 4)] if column_map.get('note') < len(row) else None
                
                # 验证必要字段
                if not date_val or not type_val or not amount_val:
                    errors.append(f'第{row_num}行: 缺少必要字段')
                    continue
                
                # 转换类型
                try:
                    if isinstance(amount_val, str):
                        amount_val = float(amount_val)
                    else:
                        amount_val = float(amount_val)
                except:
                    errors.append(f'第{row_num}行: 金额格式不正确')
                    continue
                
                # 转换日期
                if isinstance(date_val, datetime):
                    date_str = date_val.strftime('%Y-%m-%d')
                else:
                    date_str = str(date_val)
                
                # 标准化类型
                type_str = str(type_val).strip().lower()
                if '支出' in type_str or 'expense' in type_str:
                    type_str = 'expense'
                elif '收入' in type_str or 'income' in type_str:
                    type_str = 'income'
                
                category_str = str(category_val).strip() if category_val else '其他'
                note_str = str(note_val).strip() if note_val else ''
                
                # 创建交易记录
                transaction = Transaction(
                    user_id=user_id,
                    type=type_str,
                    category=category_str,
                    amount=amount_val,
                    date=date_str,
                    note=note_str
                )
                db.session.add(transaction)
                imported_count += 1
            except Exception as e:
                errors.append(f'第{row_num}行: {str(e)}')
        
        # 提交所有交易
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f'数据库保存失败: {str(e)}'}), 400
        finally:
            # 清理临时文件
            try:
                os.unlink(temp_file)
            except:
                pass
        
        return jsonify({
            'message': f'导入成功',
            'imported': imported_count,
            'errors': errors if errors else None
        }), 200
    
    except Exception as e:
        return jsonify({'error': f'处理失败: {str(e)}'}), 500

@app.route('/api/health', methods=['GET'])
def health_check():
    return jsonify({'status': 'ok'})

@app.errorhandler(404)
def not_found(error):
    return jsonify({'error': '未找到请求的资源'}), 404

@app.route('/')
def index():
    return send_from_directory('.', 'index.html')

@app.route('/<path:path>')
def serve_static(path):
    try:
        if os.path.exists(os.path.join('.', path)):
            return send_from_directory('.', path)
    except:
        pass
    return send_from_directory('.', 'index.html')

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8888))
    app.run(host='0.0.0.0', port=port, debug=False)
